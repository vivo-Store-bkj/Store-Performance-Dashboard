"""
Convert the "SALES VALUE PER TYPE" sheet inside source/dashboard_data.xlsx
into a per-month JSON snapshot (network totals by product model + per-store
breakdown), including a comparison against last month (qty & value), and
keep a manifest so the dashboard can switch between months.
Mirrors scripts/convert_xlsx_to_json.py and convert_stock_xlsx_to_json.py.

Run from the repo root: python scripts/convert_sales_type_xlsx_to_json.py

Input : source/dashboard_data.xlsx, sheet "SALES VALUE PER TYPE"
        (same file as the Performance sheet — no separate upload needed)
Output: data/sales-type-<YYYY-MM>.json   (one snapshot per month, auto-generated)
        data/sales-type-manifest.json    (list of available months, auto-generated)
Don't edit anything inside data/ by hand — it gets regenerated every run.

Sheet layout (as of Agustus 2026 template, "LINKED FROM PIVOT"):
  Row 1: title, e.g. "SALES QTY & VALUE PER TYPE - PERBANDINGAN AGUSTUS vs JULI"
  Row 3: one merged header per product model, spanning 6 columns, in order:
         [Qty bulan ini, Qty bulan lalu, Selisih Qty, Value bulan ini, Value bulan lalu, Selisih Value]
         A "TOTAL" group (same 6 columns) appears as the last group per row.
  Columns A-E: NO, AREA, ID STORE, NAMA TOKO, HEADSTORE
  Data rows start at row 5; a "TOTAL" row (network aggregate) appears at the bottom.

This positional layout (not the text labels, which change every month e.g.
"Qty Agustus" -> "Qty September") is what's parsed, so this script keeps
working as the month rolls over without edits.
"""
import openpyxl
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from validate import Validator

try:
    from zoneinfo import ZoneInfo
    JAKARTA = ZoneInfo("Asia/Jakarta")
except Exception:
    JAKARTA = None

REPO_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = REPO_ROOT / "source" / "dashboard_data.xlsx"
SHEET_NAME = "SALES VALUE PER TYPE"
DATA_DIR = REPO_ROOT / "data"
MANIFEST_PATH = DATA_DIR / "sales-type-manifest.json"

BULAN_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def format_updated_at():
    now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
    return f"{now.day} {BULAN_ID[now.month]} {now.year}, {now.strftime('%H:%M')} WIB"


def clean_area(s):
    m = re.findall(r"[A-Za-z][A-Za-z\s]*$", str(s))
    return m[-1].strip() if m else str(s)


def clean_store_name(s):
    return re.sub(r"^[A-Za-z0-9]+-", "", str(s)).strip()


def norm_model(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def find_model_groups(ws, header_row=3):
    """Each product-model group spans 6 columns starting at column F (6).
    Detect group starts by any non-empty cell in the header row."""
    groups = []
    for c in range(6, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            groups.append((c, norm_model(v)))
    return groups


def block(ws, row, start_col):
    def v(c):
        val = ws.cell(row, c).value
        return val if val is not None else 0
    return {
        "qty": v(start_col), "qty_prev": v(start_col + 1), "qty_delta": v(start_col + 2),
        "value": v(start_col + 3), "value_prev": v(start_col + 4), "value_delta": v(start_col + 5),
    }


def main():
    if not INPUT_PATH.exists():
        print(f"SKIP: {INPUT_PATH} tidak ditemukan.")
        sys.exit(0)

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"SKIP: sheet '{SHEET_NAME}' belum ada di {INPUT_PATH.name} — belum ada data sales-per-type untuk diconvert.")
        sys.exit(0)

    ws = wb[SHEET_NAME]
    groups = find_model_groups(ws)
    if not groups:
        print(f"SKIP: gagal menemukan kolom model produk di sheet {SHEET_NAME}.")
        sys.exit(0)

    total_col = next((c for c, name in groups if name.upper() == "TOTAL"), None)
    model_groups = [(c, name) for c, name in groups if name.upper() != "TOTAL"]

    DATA_DIR.mkdir(exist_ok=True)

    stores = []
    network_by_model = {}
    network_total = None

    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        if str(a).strip().upper() == "TOTAL":
            for c, name in model_groups:
                network_by_model[name] = block(ws, r, c)
            if total_col:
                network_total = block(ws, r, total_col)
            continue
        if not isinstance(a, (int, float)):
            continue
        store_name_raw = ws.cell(r, 4).value
        if not store_name_raw:
            continue

        by_type = []
        for c, name in model_groups:
            b = block(ws, r, c)
            if not (b["qty"] or b["qty_prev"] or b["value"] or b["value_prev"]):
                continue
            by_type.append({"model": name, **b})
        by_type.sort(key=lambda t: -t["value"])

        store_total = block(ws, r, total_col) if total_col else None

        stores.append({
            "store_id": ws.cell(r, 3).value,
            "store_name": clean_store_name(store_name_raw),
            "area": clean_area(ws.cell(r, 2).value),
            "headstore": ws.cell(r, 5).value,
            "total_qty": store_total["qty"] if store_total else 0,
            "total_qty_prev": store_total["qty_prev"] if store_total else 0,
            "total_value": store_total["value"] if store_total else 0,
            "total_value_prev": store_total["value_prev"] if store_total else 0,
            "by_type": by_type,
        })

    stores.sort(key=lambda s: s["store_name"])

    totals_by_type = [
        {"model": name, **b} for name, b in network_by_model.items()
    ]
    totals_by_type.sort(key=lambda t: -t["value"])

    v_st = Validator("Sales per Type")
    v_st.check(15 <= len(model_groups) <= 40, f"Jumlah tipe produk terdeteksi {len(model_groups)}, biasanya sekitar 25 — cek apakah header row 3 masih berupa grup 6-kolom per tipe")
    v_st.check(30 <= len(stores) <= 60, f"Jumlah toko yang terbaca {len(stores)}, biasanya sekitar 43")
    v_st.check(network_total is not None, "Baris TOTAL (agregat network) gak ketemu — KPI ringkasan bakal 0/kosong", level="FAIL")
    v_st.report()

    # period tag: reuse Sheet1's title if present ("DATA SALES ON <MONTH>"),
    # otherwise fall back to the current run month
    period_key = None
    period_label = None
    if "Sheet1" in wb.sheetnames or wb.worksheets[0].title != SHEET_NAME:
        main_ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
        title = str(main_ws.cell(1, 1).value or "")
        months_en = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
        if " ON " in title.upper():
            mname = title.upper().split(" ON ")[-1].strip()
            if mname in months_en:
                now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
                m_num = months_en.index(mname) + 1
                period_key = f"{now.year}-{m_num:02d}"
                period_label = f"{BULAN_ID[m_num]} {now.year}"
    if not period_key:
        now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
        period_key = f"{now.year}-{now.month:02d}"
        period_label = f"{BULAN_ID[now.month]} {now.year}"

    payload = {
        "period_key": period_key,
        "period_label": period_label,
        "updated_by": os.environ.get("UPDATED_BY", "Michael Fumar"),
        "updated_at": format_updated_at(),
        "totals_by_type": totals_by_type,
        "network_total": network_total,
        "stores": stores,
    }

    out_path = DATA_DIR / f"sales-type-{period_key}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    if MANIFEST_PATH.exists():
        manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    else:
        manifest = {"periods": []}

    existing = {p["key"]: p for p in manifest["periods"]}
    existing[period_key] = {
        "key": period_key,
        "label": period_label,
        "file": f"data/sales-type-{period_key}.json",
    }
    manifest["periods"] = sorted(existing.values(), key=lambda p: p["key"])

    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"OK: {len(stores)} toko, {len(totals_by_type)} tipe produk -> {out_path}")
    print(f"Manifest updated -> {MANIFEST_PATH} ({len(manifest['periods'])} bulan tersedia)")


if __name__ == "__main__":
    main()
