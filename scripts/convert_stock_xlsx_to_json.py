"""
Convert the daily Stock/Sales-Out/DOS Excel file (IQOO & VIVO sheets) into a
per-month JSON snapshot, and keep a manifest so the dashboard can switch
between months. Mirrors scripts/convert_xlsx_to_json.py (performance).

Run from the repo root: python scripts/convert_stock_xlsx_to_json.py

Input : source/stock_data.xlsx        (overwrite this file every time you update stock)
Output: data/stock-<YYYY-MM>.json     (one snapshot per month, auto-generated)
        data/stock-manifest.json      (list of available months, auto-generated)
Don't edit anything inside data/ by hand — it gets regenerated every run.

Source file has two sheets, "STOCK IQOO" and "STOCK VIVO", each with a
multi-row header (product model -> variant -> color -> metric). For each
store we extract:
  - the per-store TOTAL STOCK / TOTAL SALES OUT / TOTAL DOS columns (column
    position detected dynamically by header text, survives minor column
    reshuffles), and
  - a per-product-model breakdown ("by_type"): all variant/color columns
    under the same merged model header (e.g. "IQOO Z11", "VIVO Y31d") are
    summed together into one Stock/Sales Out/DOS figure per model.
"""
import openpyxl
import json
import re
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from zoneinfo import ZoneInfo
    JAKARTA = ZoneInfo("Asia/Jakarta")
except Exception:
    JAKARTA = None

REPO_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = REPO_ROOT / "source" / "stock_data.xlsx"
DATA_DIR = REPO_ROOT / "data"
MANIFEST_PATH = DATA_DIR / "stock-manifest.json"

BULAN_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]

SHEETS = [
    ("IQOO", "STOCK IQOO"),
    ("VIVO", "STOCK VIVO"),
]


def format_updated_at():
    now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
    return f"{now.day} {BULAN_ID[now.month]} {now.year}, {now.strftime('%H:%M')} WIB"


def clean_store_name(s):
    s = re.sub(r"^[A-Za-z0-9]+-", "", str(s)).strip()
    s = s.replace("仓库", "").strip()
    return s


def extract_store_id(s):
    m = re.match(r"^(\d+)-", str(s))
    return m.group(1) if m else None


def find_header_row_and_cols(ws):
    """Find the row containing 'Nama Gudang / Toko' and the TOTAL STOCK/SALES OUT/DOS columns."""
    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        v = ws.cell(r, 1).value
        if v and "nama gudang" in str(v).lower():
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Header row 'Nama Gudang / Toko' tidak ditemukan di sheet {ws.title}")

    col_stock = col_sales = col_dos = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if not v:
            continue
        vs = str(v).lower().replace("\n", " ")
        if "total" in vs and "stock" in vs:
            col_stock = c
        elif "total" in vs and "sales" in vs:
            col_sales = c
        elif "total" in vs and "dos" in vs:
            col_dos = c
    if not (col_stock and col_sales and col_dos):
        raise ValueError(f"Kolom TOTAL STOCK/SALES OUT/DOS tidak lengkap ditemukan di sheet {ws.title}")

    # data starts at the first row after header_row where column A is not empty
    data_start = header_row + 1
    while data_start <= ws.max_row and ws.cell(data_start, 1).value is None:
        data_start += 1

    # metric row (STOCK / SALES OUT / DOS / SARAN labels) is the row between
    # header_row and data_start that contains the text "STOCK"
    metric_row = None
    for r in range(header_row + 1, data_start):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and "stock" in str(v).lower():
                metric_row = r
                break
        if metric_row:
            break

    return header_row, metric_row, data_start, col_stock, col_sales, col_dos


def build_column_maps(ws, header_row, metric_row):
    """Forward-fill the merged product-model header per column, and read the
    metric label (STOCK/SALES OUT/DOS/SARAN) per column."""
    col_model = {}
    current_model = None
    for c in range(2, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            vs = str(v).lower()
            current_model = None if "total" in vs else str(v).replace("\n", " ").strip()
        col_model[c] = current_model

    col_metric = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(metric_row, c).value if metric_row else None
        col_metric[c] = str(v).strip().upper() if v else None

    return col_model, col_metric


def parse_sheet(ws):
    header_row, metric_row, data_start, col_stock, col_sales, col_dos = find_header_row_and_cols(ws)
    col_model, col_metric = build_column_maps(ws, header_row, metric_row)

    total_hari = ws.cell(1, 2).value or 30
    hari_berjalan = ws.cell(2, 2).value or 0
    sisa_hari = ws.cell(3, 2).value

    rows = []
    for r in range(data_start, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name or "grand total" in str(raw_name).lower():
            continue

        # per-model breakdown
        by_model_acc = {}
        for c in range(2, ws.max_column + 1):
            model = col_model.get(c)
            metric = col_metric.get(c)
            if not model or not metric:
                continue
            if metric not in ("STOCK", "SALES OUT"):
                continue  # ignore DOS / SARAN sub-columns, we recompute DOS ourselves
            val = ws.cell(r, c).value or 0
            acc = by_model_acc.setdefault(model, {"stock": 0, "sales_out": 0})
            if metric == "STOCK":
                acc["stock"] += val
            else:
                acc["sales_out"] += val

        by_type = []
        for model, acc in by_model_acc.items():
            if not acc["stock"] and not acc["sales_out"]:
                continue  # skip models with zero stock and zero sales at this store
            dos = round(acc["stock"] / (acc["sales_out"] / hari_berjalan), 2) if acc["sales_out"] else None
            by_type.append({
                "model": model,
                "stock": acc["stock"],
                "sales_out": acc["sales_out"],
                "dos": dos,
            })
        by_type.sort(key=lambda t: -t["stock"])

        stock = ws.cell(r, col_stock).value or 0
        sales_out = ws.cell(r, col_sales).value or 0
        dos = ws.cell(r, col_dos).value
        # source file sometimes has a sentinel DOS value (e.g. 1000) when
        # sales_out is 0 (div-by-zero workaround) — treat as "not computable"
        if not sales_out:
            dos = None

        rows.append({
            "store_id": extract_store_id(raw_name),
            "store_name": clean_store_name(raw_name),
            "stock": stock,
            "sales_out": sales_out,
            "dos": round(dos, 2) if isinstance(dos, (int, float)) else dos,
            "by_type": by_type,
        })

    return {
        "total_hari": total_hari,
        "hari_berjalan": hari_berjalan,
        "sisa_hari": sisa_hari,
        "rows": rows,
    }


def main():
    if not INPUT_PATH.exists():
        print(f"SKIP: {INPUT_PATH} tidak ditemukan — belum ada data stock untuk diconvert.")
        sys.exit(0)

    DATA_DIR.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)

    parsed = {}
    for brand, sheetname in SHEETS:
        if sheetname not in wb.sheetnames:
            print(f"WARNING: sheet '{sheetname}' tidak ditemukan, dilewati.")
            continue
        parsed[brand] = parse_sheet(wb[sheetname])

    if not parsed:
        print("ERROR: tidak ada sheet IQOO/VIVO yang bisa diparse.")
        sys.exit(1)

    hari_berjalan = next(iter(parsed.values()))["hari_berjalan"] or 1

    store_map = {}
    for brand, data in parsed.items():
        for row in data["rows"]:
            key = row["store_id"] or row["store_name"]
            store_map.setdefault(key, {"store_id": row["store_id"], "store_name": row["store_name"]})
            store_map[key][f"{brand.lower()}_stock"] = row["stock"]
            store_map[key][f"{brand.lower()}_sales_out"] = row["sales_out"]
            store_map[key][f"{brand.lower()}_dos"] = row["dos"]
            store_map[key][f"{brand.lower()}_by_type"] = row["by_type"]

    stores = []
    for key, s in store_map.items():
        iqoo_stock = s.get("iqoo_stock", 0) or 0
        vivo_stock = s.get("vivo_stock", 0) or 0
        iqoo_sales = s.get("iqoo_sales_out", 0) or 0
        vivo_sales = s.get("vivo_sales_out", 0) or 0
        total_stock = iqoo_stock + vivo_stock
        total_sales = iqoo_sales + vivo_sales
        total_dos = round(total_stock / (total_sales / hari_berjalan), 2) if total_sales else None
        s["total_stock"] = total_stock
        s["total_sales_out"] = total_sales
        s["total_dos"] = total_dos
        s.setdefault("iqoo_by_type", [])
        s.setdefault("vivo_by_type", [])
        stores.append(s)

    stores.sort(key=lambda s: s["store_name"])

    # invert the per-store by_type breakdown into a per-type view: for each
    # brand, "which stores carry model X, and how much" — this is what
    # lets the dashboard answer "tipe X ada di toko mana saja" instead of
    # only "toko Y ada tipe apa saja".
    types_by_brand = {}
    for brand in parsed.keys():
        key = f"{brand.lower()}_by_type"
        model_acc = {}
        for s in stores:
            for t in s.get(key, []):
                acc = model_acc.setdefault(t["model"], {"stock": 0, "sales_out": 0, "stores": []})
                acc["stock"] += t["stock"]
                acc["sales_out"] += t["sales_out"]
                if t["stock"] or t["sales_out"]:
                    acc["stores"].append({
                        "store_name": s["store_name"],
                        "store_id": s["store_id"],
                        "stock": t["stock"],
                        "sales_out": t["sales_out"],
                        "dos": t["dos"],
                    })
        types = []
        for model, acc in model_acc.items():
            dos = round(acc["stock"] / (acc["sales_out"] / hari_berjalan), 2) if acc["sales_out"] else None
            acc["stores"].sort(key=lambda x: -x["stock"])
            types.append({
                "model": model,
                "stock": acc["stock"],
                "sales_out": acc["sales_out"],
                "dos": dos,
                "stores": acc["stores"],
            })
        types.sort(key=lambda t: -t["stock"])
        types_by_brand[brand] = types

    now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
    period_key = f"{now.year}-{now.month:02d}"
    period_label = f"{BULAN_ID[now.month]} {now.year}"

    ref = next(iter(parsed.values()))
    payload = {
        "period": period_key,
        "period_key": period_key,
        "period_label": period_label,
        "total_hari": ref["total_hari"],
        "hari_berjalan": ref["hari_berjalan"],
        "sisa_hari": ref["sisa_hari"],
        "updated_by": os.environ.get("UPDATED_BY", "Michael Fumar"),
        "updated_at": format_updated_at(),
        "stores": stores,
        "types_by_brand": types_by_brand,
    }

    out_path = DATA_DIR / f"stock-{period_key}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    if MANIFEST_PATH.exists():
        manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    else:
        manifest = {"periods": []}

    existing = {p["key"]: p for p in manifest["periods"]}
    existing[period_key] = {
        "key": period_key,
        "label": period_label,
        "file": f"data/stock-{period_key}.json",
    }
    manifest["periods"] = sorted(existing.values(), key=lambda p: p["key"])

    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"OK: {len(stores)} toko diproses -> {out_path}")
    print(f"Manifest updated -> {MANIFEST_PATH} ({len(manifest['periods'])} bulan tersedia)")


if __name__ == "__main__":
    main()
