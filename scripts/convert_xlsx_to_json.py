"""
Convert the daily store-performance Excel file into a per-month JSON snapshot,
and keep a manifest of all months so the dashboard can switch between them.

Run from the repo root: python scripts/convert_xlsx_to_json.py

Input : source/dashboard_data.xlsx     (overwrite this file every day with your update)
Output: data/<YYYY-MM>.json            (one snapshot per month, auto-generated)
        data/manifest.json             (list of available months, auto-generated)
Don't edit anything inside data/ by hand — it gets regenerated every run.
"""
import openpyxl
import json
import re
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from validate import Validator

try:
    from zoneinfo import ZoneInfo
    JAKARTA = ZoneInfo("Asia/Jakarta")
except Exception:
    JAKARTA = None

REPO_ROOT = Path(__file__).resolve().parent.parent
INPUT_PATH = REPO_ROOT / "source" / "dashboard_data.xlsx"
DATA_DIR = REPO_ROOT / "data"
MANIFEST_PATH = DATA_DIR / "manifest.json"
HISTORY_PATH = DATA_DIR / "history.json"

BULAN_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]
MONTHS_EN = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]


def format_updated_at():
    now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
    return f"{now.day} {BULAN_ID[now.month]} {now.year}, {now.strftime('%H:%M')} WIB"


def clean_area(s):
    m = re.findall(r"[A-Za-z][A-Za-z\s]*$", str(s))
    return m[-1].strip() if m else str(s)


def clean_store_name(s):
    return re.sub(r"^[A-Za-z0-9]+-", "", str(s)).strip()


def parse_period(title):
    title = str(title)
    if " ON " in title.upper():
        return title.upper().split(" ON ")[-1].strip()
    return title.strip()


def month_number(period_name):
    p = period_name.strip().upper()
    if p in MONTHS_EN:
        return MONTHS_EN.index(p) + 1
    # fallback: kalau nama bulan gak dikenali, pakai bulan berjalan
    now = datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()
    return now.month


def find_promoter_sheet(wb):
    """Cari sheet yang punya kolom 'PROMOTOR' di salah satu baris header
    (baris 1-6) -- gak bergantung pada nama sheet, karena bisa berubah
    (contoh sekarang: 'Sheet2')."""
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, 7):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and str(v).strip().upper() == "PROMOTOR":
                    return ws
    return None


def parse_promoters(wb):
    ws = find_promoter_sheet(wb)
    if ws is None:
        return None

    def v(row, col):
        val = ws.cell(row, col).value
        return val if val is not None else 0

    promoters = []
    for r in range(5, ws.max_row + 1):
        store_id = ws.cell(r, 2).value
        promoter_name = ws.cell(r, 6).value
        if not store_id or not promoter_name:
            continue
        if not isinstance(store_id, (int, float)):
            continue  # skip baris TOTAL / bukan data promotor

        promoters.append({
            "area": clean_area(ws.cell(r, 1).value),
            "store_id": store_id,
            "headstore": ws.cell(r, 3).value,
            "store_name": clean_store_name(ws.cell(r, 5).value),
            "promoter_name": str(promoter_name).strip(),
            "mio3_target": v(r, 7),
            "mio3_ach": v(r, 8),
            "mio3_pct": v(r, 9),
            "mio3_est": v(r, 10),
            "all_target": v(r, 11),
            "all_ach": v(r, 12),
            "all_pct": v(r, 13),
            "all_est": v(r, 14),
            "mix_pct": v(r, 15),
            "value": v(r, 16),
            "asp": v(r, 17),
            "mio3_qty_prev": v(r, 18),
            "mio3_qty_curr": v(r, 19),
            "mio3_qty_gap": v(r, 20),
            "mio3_qty_pct": v(r, 21),
            "all_qty_prev": v(r, 22),
            "all_qty_curr": v(r, 23),
            "all_qty_gap": v(r, 24),
            "all_qty_pct": v(r, 25),
        })

    return promoters


def find_daily_sales_sheet(wb):
    """Cari sheet pivot 'penjualan per tanggal' -- dikenali dari isi (teks
    'Segment Price' di salah satu sel header), bukan dari nama sheet, karena
    nama sheet suka geser (Sheet2 jadi Sheet3, dst) tiap kali template diubah."""
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, 5):
            for c in range(1, min(ws.max_column, 10) + 1):
                v = ws.cell(r, c).value
                if v and str(v).strip() == "Segment Price":
                    return ws
    return None


def parse_daily_sales(wb):
    """Baca sheet pivot harian: qty & value TOTAL (3mio< + 3mio>) per toko,
    per tanggal -- ini angka penjualan HARI ITU doang (bukan kumulatif).
    Return: { 'YYYY-MM-DD': { 'store_id_str': {'qty':.., 'value':..} } }
    atau None kalau sheet-nya gak ada di file ini."""
    ws = find_daily_sales_sheet(wb)
    if ws is None:
        return None

    date_blocks = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(2, c).value
        if isinstance(v, datetime):
            date_blocks.append((c, v.date().isoformat()))
    if not date_blocks:
        return None

    result = {}
    for start_col, date_iso in date_blocks:
        day_data = {}
        for r in range(5, ws.max_row + 1):
            store_id = ws.cell(r, 2).value
            if not isinstance(store_id, (int, float)):
                continue
            qty = ws.cell(r, start_col + 4).value or 0
            value = ws.cell(r, start_col + 5).value or 0
            day_data[str(int(store_id))] = {"qty": qty, "value": value}
        result[date_iso] = day_data

    return result


def update_history_from_daily_sales(period_key, stores, daily_sales):
    """Backfill BANYAK titik history sekaligus dari sheet penjualan harian:
    kumulatifkan qty & value per toko sampai tiap tanggal, bagi target
    bulanan (dari Sheet1) buat dapetin achievement % di tanggal itu. Jauh
    lebih presisi & lengkap daripada 1 titik per upload -- 1 file bisa
    langsung ngisi puluhan hari riwayat sekaligus."""
    if HISTORY_PATH.exists():
        history = json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
    else:
        history = {"points": []}

    target_unit_by_store = {str(s["store_id"]): s["target_unit"] for s in stores}
    target_value_by_store = {str(s["store_id"]): s["target_value"] for s in stores}
    sorted_dates = sorted(daily_sales.keys())

    cum_qty = {sid: 0 for sid in target_unit_by_store}
    cum_value = {sid: 0 for sid in target_unit_by_store}
    existing_by_date = {p["date"]: p for p in history["points"]}

    for date_iso in sorted_dates:
        day = daily_sales[date_iso]
        for sid, vals in day.items():
            if sid in cum_qty:
                cum_qty[sid] += vals["qty"] or 0
                cum_value[sid] += vals["value"] or 0

        store_snapshot = {}
        total_target_unit = total_ach_unit = total_target_value = total_ach_value = 0
        for sid, target_unit in target_unit_by_store.items():
            target_value = target_value_by_store.get(sid, 0)
            ach_unit = cum_qty.get(sid, 0)
            ach_value = cum_value.get(sid, 0)
            store_snapshot[sid] = {
                "target_unit": target_unit,
                "ach_unit": ach_unit,
                "ach_pct": (ach_unit / target_unit) if target_unit else 0,
                "target_value": target_value,
                "ach_value": ach_value,
                "ach_value_pct": (ach_value / target_value) if target_value else 0,
            }
            total_target_unit += target_unit
            total_ach_unit += ach_unit
            total_target_value += target_value
            total_ach_value += ach_value

        day_num = int(date_iso.split("-")[2])
        existing_by_date[date_iso] = {
            "date": date_iso,
            "period_key": period_key,
            "to_date_day": day_num,
            "source": "daily_sales",
            "stores": store_snapshot,
            "network": {
                "target_unit": total_target_unit,
                "ach_unit": total_ach_unit,
                "ach_pct": (total_ach_unit / total_target_unit) if total_target_unit else 0,
                "target_value": total_target_value,
                "ach_value": total_ach_value,
                "ach_value_pct": (total_ach_value / total_target_value) if total_target_value else 0,
            },
        }

    history["points"] = sorted(existing_by_date.values(), key=lambda p: p["date"])
    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

    n_this_month = len([p for p in history["points"] if p["period_key"] == period_key])
    print(f"OK: history.json (dari sheet penjualan harian) -> {n_this_month} titik untuk {period_key} ({len(history['points'])} total sepanjang waktu)")


def update_history(period_key, to_date_day, stores, total_row):
    """Simpan snapshot ringan (target/achievement unit per toko) setiap kali
    script ini jalan, satu titik per HARI upload (bukan per bulan). Ini yang
    dipakai dashboard buat grafik tren dalam-bulan (mingguan/harian) --
    terpisah total dari data/<bulan>.json yang tetap cuma nyimpen versi
    TERBARU seperti biasa. File ini murni tambahan, gak pernah menghapus
    titik lama, cuma numpuk (dan menimpa titik hari yang sama kalau upload
    ulang di hari yang sama).

    Ini FALLBACK -- dipakai cuma kalau sheet penjualan harian (lihat
    parse_daily_sales) gak ketemu di file. Kalau ketemu, itu yang dipakai
    duluan karena jauh lebih presisi (lihat update_history_from_daily_sales)."""
    if HISTORY_PATH.exists():
        history = json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
    else:
        history = {"points": []}

    today = (datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()).date().isoformat()

    store_snapshot = {
        str(s["store_id"]): {
            "target_unit": s["target_unit"],
            "ach_unit": s["ach_unit"],
            "ach_pct": s["ach_pct"],
        }
        for s in stores
    }
    network_snapshot = None
    if total_row:
        network_snapshot = {
            "target_unit": total_row.get("target_unit", 0),
            "ach_unit": total_row.get("ach_unit", 0),
            "ach_pct": total_row.get("ach_pct", 0),
        }

    new_point = {
        "date": today,
        "period_key": period_key,
        "to_date_day": to_date_day,
        "source": "upload_snapshot",
        "stores": store_snapshot,
        "network": network_snapshot,
    }

    # dedup: kalau hari ini sudah pernah upload, timpa titik itu -- jangan dobel
    history["points"] = [p for p in history["points"] if p["date"] != today]
    history["points"].append(new_point)
    history["points"].sort(key=lambda p: p["date"])

    with open(HISTORY_PATH, "w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

    points_this_month = [p for p in history["points"] if p["period_key"] == period_key]
    print(f"OK: history.json (snapshot upload) -> {len(points_this_month)} titik riwayat untuk {period_key} ({len(history['points'])} total sepanjang waktu)")


def main():
    if not INPUT_PATH.exists():
        print(f"ERROR: {INPUT_PATH} tidak ditemukan. Upload file Excel ke source/dashboard_data.xlsx dulu.")
        sys.exit(1)

    DATA_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    ws = wb.worksheets[0]  # selalu ambil sheet pertama

    v_perf = Validator("Store Performance (Sheet1)")
    v_perf.expect_header(ws, 3, 1, "NO")
    v_perf.expect_header(ws, 3, 2, "AREA")
    v_perf.expect_header(ws, 3, 3, "ID STORE")
    v_perf.expect_header(ws, 3, 4, "NAMA TOKO")
    v_perf.expect_header(ws, 3, 5, "HEADSTORE")

    # --- pacing info (baris 1-2), posisinya tetap sama tiap bulan ---
    title = ws.cell(1, 1).value or ""
    total_days = ws.cell(2, 4).value or 30   # D2 = TOTAL DATE
    to_date_day = ws.cell(2, 5).value or 0   # E2 = TO DATE
    time_gone_raw = ws.cell(2, 6).value      # F2 = TIME GONE (rasio 0-1)
    time_gone = time_gone_raw if time_gone_raw is not None else (
        to_date_day / total_days if total_days else 0
    )
    v_perf.check(1 <= total_days <= 31, f"TOTAL DATE (D2) = {total_days}, di luar rentang wajar 1-31")
    v_perf.check(0 <= to_date_day <= total_days, f"TO DATE (E2) = {to_date_day}, harusnya di antara 0 dan {total_days}")
    v_perf.check(-0.01 <= time_gone <= 1.05, f"TIME GONE (F2) = {time_gone}, harusnya rasio 0-1")

    period = parse_period(title)
    m_num = month_number(period)
    year = (datetime.now(JAKARTA) if JAKARTA else datetime.utcnow()).year
    period_key = f"{year}-{m_num:02d}"
    period_label = f"{BULAN_ID[m_num]} {year}"

    def v(row, col):
        val = ws.cell(row, col).value
        return val if val is not None else 0

    # --- deteksi otomatis posisi kolom blok utama target/achievement ---
    # Struktur (per Sep 2026): kol7=target ALL TYPE(qty), kol8=target VALUE,
    # lalu blok achievement VALUE (ach/%/gap) dan blok achievement ALL TYPE
    # (ach/%/gap) muncul di posisi berbeda. Sebelumnya (s.d. Agu 2026) unit
    # yang jadi blok utama. Karena bisa ketuker lagi, kita cari lewat row4
    # (sub-header) daripada hardcode -- jadi gak langsung rusak kalau geser.
    def find_ach_blocks():
        # Cari pasangan (VALUE, %, GAP) dan (ALL TYPE, %, GAP) di antara kol 7-26.
        # Kolom % selalu tepat setelah kolom ach, GAP setelah %.
        cols = {}
        for c in range(7, 27):
            label = str(ws.cell(4, c).value or "").strip().upper()
            nxt = str(ws.cell(4, c + 1).value or "").strip().upper()
            nxt2 = str(ws.cell(4, c + 2).value or "").strip().upper()
            # blok achievement VALUE: kolomnya VALUE, diikuti % lalu GAP
            if label == "VALUE" and nxt == "%" and nxt2 == "GAP" and "ach_value" not in cols:
                cols["ach_value"] = c
            # blok achievement ALL TYPE: kolomnya ALL TYPE, diikuti % lalu GAP
            if label == "ALL TYPE" and nxt == "%" and nxt2 == "GAP" and "ach_unit" not in cols:
                cols["ach_unit"] = c
        return cols

    ACH = find_ach_blocks()
    # fallback ke posisi lama (Agu 2026) kalau deteksi gagal
    col_ach_value = ACH.get("ach_value", 18)
    col_ach_unit = ACH.get("ach_unit", 9)

    def v(row, col):
        val = ws.cell(row, col).value
        return val if val is not None else 0

    def vas_block(r):
        return {
            "acc": {"target": v(r, 55), "ach": v(r, 56), "pct": v(r, 57)},
            "qoala": {"target": v(r, 58), "ach": v(r, 59), "pct": v(r, 60)},
            "bca_insurance": {"target": v(r, 61), "ach": v(r, 62), "pct": v(r, 63)},
            "indosat": {"target": v(r, 64), "ach": v(r, 65), "pct": v(r, 66)},
            "telkomsel": {"target": v(r, 67), "ach": v(r, 68), "pct": v(r, 69)},
        }

    def ach_block(r):
        """Ambil achievement value & unit dari posisi yang sudah dideteksi."""
        return {
            "ach_value": v(r, col_ach_value),
            "ach_value_pct": v(r, col_ach_value + 1),
            "gap_value": v(r, col_ach_value + 2),
            "ach_unit": v(r, col_ach_unit),
            "ach_pct": v(r, col_ach_unit + 1),
            "gap_unit": v(r, col_ach_unit + 2),
        }

    stores = []
    total_row = None

    for r in range(5, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        if str(a).strip().upper() == "TOTAL":
            total_row = {
                "target_unit": v(r, 7), "target_value": v(r, 8),
                **ach_block(r),
                "mio3_unit": v(r, 13), "mio3_pct": v(r, 14),
                "iqoo_unit": v(r, 15), "iqoo_pct": v(r, 16),
                "growth_all_prev": v(r, 35), "growth_all_curr": v(r, 36),
                "growth_all_gap": v(r, 37), "growth_all_pct": v(r, 38),
                "growth_value_prev": v(r, 47), "growth_value_curr": v(r, 48),
                "growth_value_gap": v(r, 49), "growth_value_pct": v(r, 50),
                "vas": vas_block(r),
            }
            continue
        if not isinstance(a, (int, float)):
            continue  # skip baris kosong/footer lain

        stores.append({
            "no": v(r, 1),
            "area": clean_area(v(r, 2)),
            "area_raw": v(r, 2),
            "store_id": v(r, 3),
            "store_name": clean_store_name(v(r, 4)),
            "manager": v(r, 5),
            "jml_pc": v(r, 6),
            "target_unit": v(r, 7),
            "target_value": v(r, 8),
            **ach_block(r),
            "mio3_unit": v(r, 13),
            "mio3_pct": v(r, 14),
            "iqoo_unit": v(r, 15),
            "iqoo_pct": v(r, 16),
            "growth_all_prev": v(r, 35),
            "growth_all_curr": v(r, 36),
            "growth_all_gap": v(r, 37),
            "growth_all_pct": v(r, 38),
            "growth_value_prev": v(r, 47),
            "growth_value_curr": v(r, 48),
            "growth_value_gap": v(r, 49),
            "growth_value_pct": v(r, 50),
            "vas": vas_block(r),
        })

    v_perf.check(30 <= len(stores) <= 60, f"Jumlah toko yang terbaca {len(stores)}, biasanya sekitar 43 — cek apakah ada baris yang kelewat/kebaca ganda", level="WARN" if 20 <= len(stores) <= 70 else "FAIL")
    v_perf.check(total_row is not None, "Baris TOTAL tidak ketemu di Sheet1 — KPI network-wide bakal kosong", level="FAIL")

    promoters = parse_promoters(wb)
    if promoters is not None:
        store_ids = {s["store_id"] for s in stores}
        orphan_ids = {p["store_id"] for p in promoters} - store_ids
        v_perf.check(
            len(orphan_ids) == 0,
            f"{len(orphan_ids)} store_id di sheet promotor gak ketemu di Sheet1 (contoh: {sorted(orphan_ids)[:5]}) — promotor itu gak akan muncul di dashboard toko manapun",
        )
        v_perf.check(len(promoters) > 0, "Sheet promotor ketemu tapi isinya 0 baris — cek apakah kolom ID Store/PROMOTOR kosong semua")

    daily_sales_preview = parse_daily_sales(wb)
    if daily_sales_preview is not None:
        sum_qty = sum(
            (vals.get("qty") or 0)
            for day in daily_sales_preview.values()
            for vals in day.values()
        )
        ref_ach = total_row.get("ach_unit", 0) if total_row else 0
        v_perf.check(
            sum_qty == ref_ach,
            f"Total qty dari sheet penjualan harian ({sum_qty}) beda dari baris TOTAL Sheet1 ({ref_ach}) — kemungkinan rentang tanggalnya gak lengkap/dobel",
        )
        v_perf.check(len(daily_sales_preview) > 0, "Sheet penjualan harian ketemu tapi 0 blok tanggal terdeteksi")

    v_perf.report()

    data = {
        "period": period,
        "period_key": period_key,
        "period_label": period_label,
        "to_date_day": to_date_day,
        "total_days": total_days,
        "time_gone": time_gone,
        "updated_by": os.environ.get("UPDATED_BY", "Michael Fumar"),
        "updated_at": format_updated_at(),
        "stores": stores,
        "total": total_row,
        "promoters": promoters,
    }

    out_path = DATA_DIR / f"{period_key}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # --- update manifest (daftar semua bulan yang tersedia) ---
    if MANIFEST_PATH.exists():
        manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    else:
        manifest = {"periods": []}

    existing = {p["key"]: p for p in manifest["periods"]}
    existing[period_key] = {
        "key": period_key,
        "label": period_label,
        "file": f"data/{period_key}.json",
    }
    manifest["periods"] = sorted(existing.values(), key=lambda p: p["key"])

    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"OK: {len(stores)} toko diproses -> {out_path}")
    print(f"Manifest updated -> {MANIFEST_PATH} ({len(manifest['periods'])} bulan tersedia)")

    daily_sales = parse_daily_sales(wb)
    if daily_sales:
        update_history_from_daily_sales(period_key, stores, daily_sales)
    else:
        update_history(period_key, to_date_day, stores, total_row)


if __name__ == "__main__":
    main()
